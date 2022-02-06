#!/usr/bin/env python3

import datetime
from sys import api_version
from airflow import DAG
from airflow.operators.python_operator import PythonOperator
from airflow.operators.docker_operator import DockerOperator
from airflow.utils.dates import days_ago
from docker.types import Mount
from openpyxl import load_workbook
import pandas as pd
import requests
import os

# ----------------------------------------------------------------------- #

def extract_xlsx_sheet(file_location,sheet,output_filelocation):
    wb = load_workbook(file_location)
    sheets = wb.sheetnames

    for s in sheets:
        if s != sheet:
            sheet_name = wb.get_sheet_by_name(s)
            wb.remove_sheet(sheet_name)
    
    wb.save(output_filelocation)

def process_xlsx_sheet_data(xlsx_file_location):
    df = pd.read_excel(xlsx_file_location)

    print("DEBUG: Fazendo UNION dos Dataframes da tabela vendas_derivados_oleo")

    JANEIRO = df[['COMBUSTÍVEL', 'ANO', 'ESTADO', 'Jan']].rename(columns={"Jan": "VOLUME"}).astype({'ANO': 'str'})
    FEVEREIRO = df[['COMBUSTÍVEL', 'ANO', 'ESTADO', 'Fev']].rename(columns={"Fev": "VOLUME"}).astype({'ANO': 'str'})
    MARCO = df[['COMBUSTÍVEL', 'ANO', 'ESTADO', 'Mar']].rename(columns={"Mar": "VOLUME"}).astype({'ANO': 'str'})
    ABRIL = df[['COMBUSTÍVEL', 'ANO', 'ESTADO', 'Abr']].rename(columns={"Abr": "VOLUME"}).astype({'ANO': 'str'})
    MAIO = df[['COMBUSTÍVEL', 'ANO', 'ESTADO', 'Mai']].rename(columns={"Mai": "VOLUME"}).astype({'ANO': 'str'})
    JUNHO = df[['COMBUSTÍVEL', 'ANO', 'ESTADO', 'Jun']].rename(columns={"Jun": "VOLUME"}).astype({'ANO': 'str'})
    JULHO = df[['COMBUSTÍVEL', 'ANO', 'ESTADO', 'Jul']].rename(columns={"Jul": "VOLUME"}).astype({'ANO': 'str'})
    AGOSTO = df[['COMBUSTÍVEL', 'ANO', 'ESTADO', 'Ago']].rename(columns={"Ago": "VOLUME"}).astype({'ANO': 'str'})
    SETEMBRO = df[['COMBUSTÍVEL', 'ANO', 'ESTADO', 'Set']].rename(columns={"Set": "VOLUME"}).astype({'ANO': 'str'})
    OUTUBRO = df[['COMBUSTÍVEL', 'ANO', 'ESTADO', 'Out']].rename(columns={"Out": "VOLUME"}).astype({'ANO': 'str'})
    NOVEMBRO = df[['COMBUSTÍVEL', 'ANO', 'ESTADO', 'Nov']].rename(columns={"Nov": "VOLUME"}).astype({'ANO': 'str'})
    DEZEMBRO = df[['COMBUSTÍVEL', 'ANO', 'ESTADO', 'Dez']].rename(columns={"Dez": "VOLUME"}).astype({'ANO': 'str'})

    JANEIRO['year_month'] = JANEIRO['ANO'] + '-' + '01'
    FEVEREIRO['year_month'] = FEVEREIRO['ANO'] + '-' + '02'
    MARCO['year_month'] = MARCO['ANO'] + '-' + '03'
    ABRIL['year_month'] = ABRIL['ANO'] + '-' + '04'
    MAIO['year_month'] = MAIO['ANO'] + '-' + '05'
    JUNHO['year_month'] = JUNHO['ANO'] + '-' + '06'
    JULHO['year_month'] = JULHO['ANO'] + '-' + '07'
    AGOSTO['year_month'] = AGOSTO['ANO'] + '-' + '08'
    SETEMBRO['year_month'] = SETEMBRO['ANO'] + '-' + '09'
    OUTUBRO['year_month'] = OUTUBRO['ANO'] + '-' + '10'
    NOVEMBRO['year_month'] = NOVEMBRO['ANO'] + '-' + '11'
    DEZEMBRO['year_month'] = DEZEMBRO['ANO'] + '-' + '12'

    df = (
                    pd.concat([
                                JANEIRO,
                                FEVEREIRO,
                                MARCO,
                                ABRIL,
                                MAIO,
                                JUNHO,
                                JULHO,
                                AGOSTO,
                                SETEMBRO,
                                OUTUBRO,
                                NOVEMBRO,
                                DEZEMBRO
                            ]
                        )
                        .astype({
                                    'VOLUME': 'float',
                                    'year_month': 'datetime64',
                                })
                        .rename(
                            columns={
                                'COMBUSTÍVEL': 'product',
                                'ESTADO': 'uf',
                                'REGIAO': 'unit',
                                'VOLUME': 'volume'
                            }
                        )
                        .drop(columns=['ANO'])
                )

    state_uf_dict = [
        ['ACRE','AC'],
        ['ALAGOAS','AL'],
        ['AMAPÁ','AP'],
        ['AMAZONAS','AM'],
        ['BAHIA','BA'],
        ['CEARÁ','CE'],
        ['DISTRITO FEDERAL','DF'],
        ['ESPÍRITO SANTO','ES'],
        ['GOIÁS','GO'],
        ['MARANHÃO','MA'],
        ['MATO GROSSO DO SUL','MS'],
        ['MATO GROSSO','MT'],
        ['MINAS GERAIS','MG'],
        ['PARANÁ','PR'],
        ['PARAÍBA','PB'],
        ['PARÁ','PA'],
        ['PERNAMBUCO','PE'],
        ['PIAUÍ','PI'],
        ['RIO DE JANEIRO','RJ'],
        ['RIO GRANDE DO NORTE','RN'],
        ['RIO GRANDE DO SUL','RS'],
        ['RONDÔNIA','RO'],
        ['RORAIMA','RR'],
        ['SANTA CATARINA','SC'],
        ['SERGIPE','SE'],
        ['SÃO PAULO','SP'],
        ['TOCANTINS','TO']
    ]

    for x in state_uf_dict:
        df['uf'] = df['uf'].replace(x[0],x[1])

    print("DEBUG: Adicionando data de criação")                    
    df['created_at'] = pd.to_datetime('today')

    return df

def check_source_with_output(source_xlsx,output_parquet,collumn_name_source,collumn_name_output):
    df1 = pd.read_excel(source_xlsx)
    df2 = pd.read_parquet(output_parquet)
    df1[collumn_name_source].equals(df2[collumn_name_output])


# ----------------------------------------------------------------------- #

def get_xls_from_github():
    print("DEBUG: Fazendo download do xls")
    file_url='https://github.com/raizen-analytics/data-engineering-test/raw/master/assets/vendas-combustiveis-m3.xls'
    file_name='/raw_zone/dados_brutos.xls'
    request = requests.get(file_url, allow_redirects=True)
    open(file_name, 'wb').write(request.content)

def create_oil_derivate_xlsx_file():
    print("DEBUG: Criando o arquivo /stage_zone/oil_derivates.xlsx")
    extract_xlsx_sheet('/raw_zone/dados_brutos.xlsx','DPCache_m3','/stage_zone/oil_derivates.xlsx')

def create_diesel_xlsx_file():
    print("DEBUG: Criando o arquivo /stage_zone/diesel.xlsx")
    extract_xlsx_sheet('/raw_zone/dados_brutos.xlsx','DPCache_m3_2','/stage_zone/diesel.xlsx')

def process_and_create_parquet_oil_derivate():
    print("DEBUG: Carregando dados do xlsx oil_derivates.xlsx no Pandas Dataframe")
    df = process_xlsx_sheet_data('/stage_zone/oil_derivates.xlsx')
    print("DEBUG: Salvando dados como parquet")
    df.to_parquet('/refined_zone/oil_derivates.parquet', compression='snappy', partition_cols=['uf','product'])

def process_and_create_parquet_diesel_derivate():
    print("DEBUG: Carregando dados do xlsx diesel.xlsx no Pandas Dataframe")
    df = process_xlsx_sheet_data('/stage_zone/diesel.xlsx')
    print("DEBUG: Salvando dados como parquet")
    df.to_parquet('/refined_zone/diesel.parquet', compression='snappy', partition_cols=['uf','product'])

def check_diesel_product_collumn():
    check_source_with_output('/stage_zone/diesel.xlsx','/refined_zone/diesel.parquet','COMBUSTÍVEL','product')

def check_oil_derivate_product_collumn():
    check_source_with_output('/stage_zone/oil_derivates.xlsx','/refined_zone/oil_derivates.parquet','COMBUSTÍVEL','product')


# ----------------------------------------------------------------------- #

default_args = {
    'owner': 'Cleiton Moura Loura',
    'depends_on_past': False,
    'email': ['cleitonmloura@gmail.com'],
    'email_on_failure': False,
    'email_on_retry': False,
    'retries': 1,
    'retry_delay': datetime.timedelta(minutes=1),
}

with DAG(
    'raizen_challenge',
    default_args=default_args,
    description='Realiza processo ETL no arquivo xls do Desafio Raízen',
    schedule_interval=datetime.timedelta(days=1),
    start_date=days_ago(2),
    max_active_runs=1,
    tags=['xls','raizen','challenge','etl'],
) as dag:

    get_xls_from_source_location = PythonOperator(
        task_id="get_xls_from_source_location",
        python_callable=get_xls_from_github
    )

    current_directory = os.environ['DATALAKE_PATH']
    file_name = 'dados_brutos.xls'

    convert_xls_to_xlsx = DockerOperator(
        task_id='convert_xls_to_xlsx',
        image='ipunktbs/docker-libreoffice-headless:latest',
        api_version='auto',
        auto_remove=True,
        command='libreoffice --invisible --headless --convert-to xlsx --outdir /tmp "{0}"'.format("/tmp/" + file_name),
        docker_url="unix:///var/run/docker.sock",
        network_mode="bridge",
        mount_tmp_dir=False,
        mounts=[Mount(source=current_directory + '/raw_zone/',target="/tmp", type="bind")]
    )

    create_oil_derivate_xlsx = virtualenv_task = PythonOperator(
        task_id="create_oil_derivate_xlsx",
        python_callable=create_oil_derivate_xlsx_file
    )

    create_diesel_derivate_xlsx = PythonOperator(
        task_id="create_diesel_derivate_xlsx",
        python_callable=create_diesel_xlsx_file
    )

    process_and_store_parquet_oil_derivate = PythonOperator(
        task_id="process_and_store_parquet_oil_derivate",
        python_callable=process_and_create_parquet_oil_derivate
    )

    process_and_store_parquet_diesel = PythonOperator(
        task_id="process_and_store_parquet_diesel",
        python_callable=process_and_create_parquet_diesel_derivate
    )

    check_diesel_product = PythonOperator(
        task_id="check_diesel_product",
        python_callable=check_diesel_product_collumn
    )

    check_oil_derivate_product = PythonOperator(
        task_id="check_oil_derivate_product",
        python_callable=check_oil_derivate_product_collumn
    )

    get_xls_from_source_location >> convert_xls_to_xlsx 
    convert_xls_to_xlsx >> create_diesel_derivate_xlsx
    convert_xls_to_xlsx >> create_oil_derivate_xlsx
    create_oil_derivate_xlsx >> process_and_store_parquet_oil_derivate
    create_diesel_derivate_xlsx >> process_and_store_parquet_diesel
    process_and_store_parquet_oil_derivate >> check_oil_derivate_product
    process_and_store_parquet_diesel >> check_diesel_product